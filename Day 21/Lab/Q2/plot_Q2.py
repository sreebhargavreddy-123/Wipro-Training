import os
import pandas as pd
from openpyxl import Workbook, load_workbook

INPUT_FILE = "sales_data.xlsx"
SHEET_NAME = "2026"

print("Working Directory:", os.getcwd())
print("-" * 50)

# =================================================
# STEP 1: CREATE VALID EXCEL FILE IF NEEDED
# =================================================
def create_valid_excel():
    print("Creating a valid Excel file...")

    wb = Workbook()
    sheet = wb.active
    sheet.title = SHEET_NAME

    sheet.append(["Product", "Quantity", "Price"])
    sheet.append(["A", 10, 50])
    sheet.append(["B", 20, 30])
    sheet.append(["C", 15, 40])

    wb.save(INPUT_FILE)
    print("Valid sales_data.xlsx created")


# If file doesn't exist OR is invalid → recreate
try:
    load_workbook(INPUT_FILE)
except Exception:
    create_valid_excel()

print("-" * 50)

# =================================================
# PART 1: USING PANDAS
# =================================================
print("Processing using Pandas...")

try:
    df = pd.read_excel(
        INPUT_FILE,
        sheet_name=SHEET_NAME,
        engine="openpyxl"
    )

    df["Total"] = df["Quantity"] * df["Price"]
    df.to_excel("sales_summary_pandas.xlsx", index=False)

    print("sales_summary_pandas.xlsx created successfully")

except Exception as e:
    print("Pandas processing failed:", e)

print("-" * 50)

# =================================================
# PART 2: USING OPENPYXL (NO PANDAS)
# =================================================
print("Processing using OpenPyXL...")

try:
    wb = load_workbook(INPUT_FILE)
    sheet = wb[SHEET_NAME]

    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.title = "Sales Summary"

    new_sheet.append(["Product", "Quantity", "Price", "Total"])

    for row in sheet.iter_rows(min_row=2, values_only=True):
        product, quantity, price = row
        new_sheet.append([product, quantity, price, quantity * price])

    new_wb.save("sales_summary_openpyxl.xlsx")

    print("sales_summary_openpyxl.xlsx created successfully")

except Exception as e:
    print("OpenPyXL processing failed:", e)

print("-" * 50)
print("Processing completed successfully.")